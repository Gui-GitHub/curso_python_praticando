from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

# Caminho do arquivo Excel
PASTA_RAIZ = Path(__file__).parent
ARQUIVO_EXCEL = PASTA_RAIZ / 'exemplo_openpyxl.xlsx'

# 1. Criando um novo arquivo Excel
wb = Workbook()
ws = wb.active
ws.title = "Funcionários"

# 2. Adicionando cabeçalho com estilo
cabecalho = ['Nome', 'Cargo', 'Idade']
ws.append(cabecalho)

# Estilo: fundo amarelo, texto vermelho, negrito
for cell in ws[1]:
    cell.font = Font(bold=True, color="FF0000")  # Vermelho e negrito
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarelo

# 3. Adicionando dados fictícios
dados = [
    ['Alice', 'Analista', 30],
    ['Bob', 'Desenvolvedor', 25],
    ['Carol', 'Gerente', 40],
    ['David', 'Designer', 28],
]
for linha in dados:
    ws.append(linha)

# 4. Salvando o arquivo Excel
wb.save(ARQUIVO_EXCEL)

# 5. Lendo e exibindo o conteúdo do Excel
print("Lendo dados do Excel:")
wb_lido = load_workbook(ARQUIVO_EXCEL)
ws_lido = wb_lido.active

for linha in ws_lido.iter_rows(values_only=True):
    print(linha)
